from __future__ import annotations

from pathlib import Path
import sys
import xml.etree.ElementTree as ET
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _local_name(tag: str) -> str:
    """Return the local name of an XML tag, ignoring namespace."""
    return tag.split('}', 1)[-1] if '}' in tag else tag


def _find_child_text(parent: ET.Element, child_name: str, default: str = "") -> str:
    """Find direct child text by local tag name."""
    for child in parent:
        if _local_name(child.tag) == child_name:
            return (child.text or "").strip()
    return default


def _collect_error_details(report_element: ET.Element | None) -> str:
    """Collect all error details under <report>."""
    if report_element is None:
        return ""

    details: list[str] = []
    for element_report in report_element:
        if _local_name(element_report.tag) != "elementReport":
            continue

        error_code = _find_child_text(element_report, "operationErrorCode")
        error_detail = (
            _find_child_text(element_report, "operationErrorDetail")
            or _find_child_text(element_report, "operationDetail")
        )

        if error_code and error_detail:
            details.append(f"{error_code}: {error_detail}")
        elif error_detail:
            details.append(error_detail)
        elif error_code:
            details.append(error_code)

    return "\n".join(details)


def parse_response_xml(xml_path: str | Path) -> list[dict[str, str]]:
    """Parse a EUDAMED response XML file into rows for Excel output.

    Returns a list of dictionaries with keys:
    - EUDAMED Response ID
    - UDI
    - ResponseCode
    - ErrorDetail
    """
    xml_path = Path(xml_path)
    tree = ET.parse(xml_path)
    root = tree.getroot()

    rows: list[dict[str, str]] = []

    for element in root.iter():
        if _local_name(element.tag) != "responseEntity":
            continue

        udi = _find_child_text(element, "entityCode")
        response_code = _find_child_text(element, "responseCode")

        report = None
        for child in element:
            if _local_name(child.tag) == "report":
                report = child
                break

        error_detail = _collect_error_details(report)

        rows.append(
            {
                "EUDAMED Response ID": xml_path.stem,
                "UDI": udi,
                "ResponseCode": response_code,
                "ErrorDetail": error_detail,
            }
        )

    return rows


def parse_multiple_response_xml(xml_paths: Iterable[str | Path]) -> list[dict[str, str]]:
    """Parse multiple EUDAMED response XML files and merge rows."""
    all_rows: list[dict[str, str]] = []
    for xml_path in xml_paths:
        all_rows.extend(parse_response_xml(xml_path))
    return all_rows


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            for line in value.splitlines() or [""]:
                max_len = max(max_len, len(line))
        adjusted = min(max_len + 2, 100)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, adjusted)


def rows_to_excel(rows: Iterable[dict[str, str]], output_path: str | Path) -> Path:
    """Write parsed rows to an Excel file."""
    output_path = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "List"

    headers = ["EUDAMED Response ID", "UDI", "ResponseCode", "ErrorDetail"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([
            row.get("EUDAMED Response ID", ""),
            row.get("UDI", ""),
            row.get("ResponseCode", ""),
            row.get("ErrorDetail", ""),
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    _autosize_columns(ws)
    ws.column_dimensions["A"].width = min(max(ws.column_dimensions["A"].width, 24), 40)
    ws.column_dimensions["D"].width = min(max(ws.column_dimensions["D"].width, 40), 120)

    wb.save(output_path)
    return output_path


def build_output_path_from_first_xml(xml_paths: Iterable[str | Path]) -> Path:
    """Create output Excel path: <first_xml_stem>_List.xlsx"""
    xml_paths = list(xml_paths)
    if not xml_paths:
        raise ValueError("xml_paths cannot be empty")
    first_xml = Path(xml_paths[0])
    return first_xml.with_name(f"{first_xml.stem}_List.xlsx")


def xml_to_excel(xml_path: str | Path, output_path: str | Path | None = None) -> Path:
    """Convert a single XML file to Excel."""
    xml_path = Path(xml_path)
    if output_path is None:
        output_path = build_output_path_from_first_xml([xml_path])

    rows = parse_response_xml(xml_path)
    return rows_to_excel(rows, output_path)


def xmls_to_merged_excel(xml_paths: Iterable[str | Path], output_path: str | Path | None = None) -> Path:
    """Convert multiple XML files into one merged Excel file.

    Output filename uses the first XML filename as prefix.
    """
    xml_paths = [Path(p) for p in xml_paths]
    if not xml_paths:
        raise ValueError("Please provide at least one XML file.")

    if output_path is None:
        output_path = build_output_path_from_first_xml(xml_paths)

    rows = parse_multiple_response_xml(xml_paths)
    return rows_to_excel(rows, output_path)


def main(argv: list[str] | None = None) -> int:
    argv = argv or sys.argv[1:]
    if not argv:
        print("Usage: python eudamed_response_xml_to_excel.py <response1.xml> [response2.xml ...]")
        return 1

    try:
        output = xmls_to_merged_excel(argv)
        print(f"OK: merged {len(argv)} file(s) -> {output}")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
